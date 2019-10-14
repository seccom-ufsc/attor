'''Module for matching Sympla check-ins with UFSC's classes.'''
from __future__ import annotations

from dataclasses import is_dataclass
from datetime import date as Date
from getpass import getpass
from pathlib import Path
from sys import argv
from textwrap import dedent
from typing import (
    Any,
    Iterable,
    List,
    NamedTuple,
    Optional,
    Sequence,
    Tuple,
)
import csv
import dataclasses

from carl import command, REQUIRED
from carl.carl import Command
from cagrex import CAGR
from cagrex.cagr import Student
from openpyxl import load_workbook
from openpyxl.cell.read_only import ReadOnlyCell
from zipfile import BadZipfile


class TypeInferError(Exception):
    pass


class Class(NamedTuple):
    subject_id: str
    class_id: str
    semester: str


class Attender(NamedTuple):
    name: str
    attended: bool = True
    student_id: Optional[str] = None
    cpf: Optional[str] = None
    classes: List[str] = []


class Ticket(NamedTuple):
    number: int
    ticket_id: str
    name: str
    surname: str
    ticket_type: str
    value: str
    order_date: Date
    order_id: str
    email: str
    state: str
    checked_in: str
    checkin_date: Optional[Date]
    discount_code: str
    pay_method: str
    pdv: str
    cpf: str
    student_id: str
    classes: List[str]

    @staticmethod
    def from_row(row) -> Ticket:
        row = [
            cell.value if cell.value else ''
            for cell in row
        ]
        return Ticket(
            int(row[0]),
            row[1],
            row[2],
            row[3],
            row[4],
            row[5],
            Date.fromisoformat(row[6].split(' ')[0]),
            row[7],
            row[8],
            row[9],
            row[10],
            Date.fromisoformat(row[11].split(' ')[0]) if row[11] else None,
            row[12],
            row[13],
            row[14],
            row[15],
            row[16],
            row[17].split(','),
        )


def iter_as_tickets(
    sheet_range: Iterable[Tuple[ReadOnlyCell]]
) -> Iterable[Ticket]:
    for row in sheet_range:
        if all(cell.value is None for cell in row):
            return

        yield Ticket.from_row(row)


def strip_join(c: Tuple[str, str]) -> str:
    return ' '.join(s.strip() for s in list(c)).strip()


def retrieve_attenders(source: Path) -> List[Attender]:
    wb = load_workbook(filename=source.resolve(), read_only=True)
    sheet = wb.active

    row_iter = sheet.iter_rows(min_row=9, min_col=1, max_col=18)
    rows = [
        (strip_join((ticket.name, ticket.surname)).title(),
         ticket.checked_in == 'Sim',
         ticket.student_id, )
        for ticket in iter_as_tickets(row_iter)
        if ticket.checked_in == 'Sim'
    ]

    return [Attender(*row) for row in rows]


def attenders_from_class(source: Path, class_file: Path) -> List[Attender]:
    attenders = data_from_csv(source)
    students = data_from_csv(class_file)
    student_ids = [student.student_id for student in students]

    return [
        attender for attender in attenders
        if attender.student_id in student_ids
    ]


def infer_type_from_fields(fieldnames: Sequence[str]):
    for _type in [Attender, Student]:
        if fieldnames == list(_type._fields):
            return _type
    raise TypeInferError(f'Could infer from fieldnames: {fieldnames}')


def data_from_csv(path: Path):
    with open(path) as f:
        reader = csv.DictReader(f)
        _type = infer_type_from_fields(reader.fieldnames)
        return [_type(**d) for d in reader]


def save_into_csv(data: List[Any], output: Path):
    if is_dataclass(data[0]):
        fields = [field.name for field in dataclasses.fields(data[0])]
        values_dict = (dataclasses.asdict(value) for value in data)
    else:
        fields = type(data[0])._fields
        values_dict = (value._asdict() for value in data)

    with open(output, 'w') as out:
        writer = csv.DictWriter(
            out,
            [f.replace('_', '') for f in fields],
            quoting=csv.QUOTE_NONE,
        )
        writer.writeheader()
        writer.writerows(sorted([
            {key.replace('_', ''): value for key, value in entry.items()}
            for entry in values_dict
        ], key=lambda x: x['name']))


@command
def main(subcommand: Command = REQUIRED):
    pass


@main.subcommand
def convert(source: Path, output: Path):
    try:
        data = retrieve_attenders(Path(source))
    except BadZipfile:
        data = data_from_csv(Path(source))
    save_into_csv(data, output)


@main.subcommand
def filter(attenders_csv: Path, class_members_csv: Path, output: Path):
    attenders = data_from_csv(attenders_csv)
    classmembers = data_from_csv(class_members_csv)

    filtered = [
        attender
        for attender in attenders
        if any(
            attender.student_id == member.student_id
            for member in classmembers
        )
    ]

    save_into_csv(filtered, output)


@main.subcommand
def fetch_members(
    subject_id: str,
    class_id: str,
    semester: str,
    output_dir: Path
):
    output_dir = Path(output_dir)
    cagr = CAGR()

    print('**CAGR Login**')
    cagr.login(input('UFSC ID: '), getpass('Password: '))

    students = cagr.students_from_class(subject_id, class_id, semester)

    output_dir = output_dir / semester / subject_id
    if not output_dir.exists():
        output_dir.mkdir(parents=True)

    save_into_csv(students, output_dir / f'{class_id}.csv')


if __name__ == '__main__':
    main.run()
