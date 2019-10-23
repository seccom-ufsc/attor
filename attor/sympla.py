'''Module to deal with sympla attendance spreadsheets.'''
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date as Date, datetime as DateTime, time as Time
from pathlib import Path
from typing import Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.cell.read_only import ReadOnlyCell


@dataclass(frozen=True)
class Attender:
    name: str
    attended: bool = True
    student_id: Optional[str] = None
    cpf: Optional[str] = None
    classes: List[str] = field(default_factory=list)


@dataclass(frozen=True)
class Ticket:
    number: int
    ticket_id: str
    name: str
    surname: str
    ticket_type: str
    value: str
    order_date: Date
    order_id: str
    email: str
    state: str
    checked_in: bool
    checkin_date: Optional[DateTime]
    discount_code: str
    pay_method: str
    pdv: str
    cpf: str
    student_id: Optional[str]
    classes: List[str]

    @staticmethod
    def from_row(row) -> Ticket:
        row = [
            cell.value if cell.value else ''
            for cell in row
        ]
        return Ticket(
            number=int(row[0]),
            ticket_id=row[1],
            name=row[2],
            surname=row[3],
            ticket_type=row[4],
            value=row[5],
            order_date=Date.fromisoformat(row[6].split(' ')[0]),
            order_id=row[7],
            email=row[8],
            state=row[9],
            checked_in=row[10] == 'Sim',
            checkin_date=(
                DateTime.fromisoformat(row[11]) if row[11] else None
            ),
            discount_code=row[12],
            pay_method=row[13],
            pdv=row[14],
            cpf=str(row[15]),
            student_id=str(int(row[16])) if row[16] else None,
            classes=row[17].split(','),
        )


@dataclass
class Sheet:
    name: str
    date: Date
    start: Time
    end: Time
    tickets: List[Ticket]

    @staticmethod
    def load(path: Path, name: str = None) -> Sheet:
        wb = load_workbook(filename=path.resolve(), read_only=True)
        sheet = wb.active

        tickets = extract_tickets(sheet)

        start, end = map(lambda x: x[0].value, sheet['A6:A7'])
        date = start.date

        return Sheet(
            name=name or sheet.title,
            date=date(),
            start=start.time(),
            end=end.time(),
            tickets=tickets,
        )


def iter_as_tickets(
    sheet_range: Iterable[Tuple[ReadOnlyCell]]
) -> Iterable[Ticket]:
    for row in sheet_range:
        if all(cell.value is None for cell in row):
            return

        yield Ticket.from_row(row)


def strip_join(c: Tuple[str, str]) -> str:
    return ' '.join(s.strip() for s in list(c)).strip()


def extract_tickets(sheet) -> List[Ticket]:
    row_iter = sheet.iter_rows(min_row=9, min_col=1, max_col=18)
    return list(iter_as_tickets(row_iter))
